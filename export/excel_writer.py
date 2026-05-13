"""
Xuất dữ liệu theo định dạng "Data Entry" - Bảng kê chi tiết nhập hàng.
Mỗi dòng là một line-item, kèm theo thông tin hóa đơn ở đầu dòng.
"""

import logging
from pathlib import Path
from typing import List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class ExcelWriter:
    def __init__(self):
        self.header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid") # Màu xanh lá nhạt
        self.header_font = Font(bold=True)
        self.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def write_results(self, results: List[Any], output_path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Dữ Liệu Nhập Hóa Đơn"

        # Định nghĩa các cột chuẩn cho "Data Entry"
        # Chúng ta sẽ gom các trường quan trọng nhất vào một bảng phẳng
        headers = [
            "STT", "Tên File", "Loại Chứng Từ", "Số Chứng Từ", "Ngày", 
            "Người Bán/Người Gửi", "Người Mua/Người Nhận", 
            "Tên Sản Phẩm/Mô Tả", "Số Lượng", "Đơn Vị", "Đơn Giá", "Thành Tiền", 
            "Tiền Tệ", "Incoterms", "Độ Tin Cậy (%)"
        ]
        
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        stt = 1
        for res in results:
            data = res.data
            # Trích xuất thông tin chung (header info)
            inv_no = self._get_val(data, ["invoice_no", "packing_list_no", "bl_no", "certificate_no"])
            date = self._get_val(data, ["date"])
            seller = self._get_val(data, ["seller_name", "shipper", "exporter"])
            buyer = self._get_val(data, ["buyer_name", "consignee", "importer"])
            currency = self._get_val(data, ["currency"])
            incoterms = self._get_val(data, ["incoterms"])

            # Nếu có danh sách mặt hàng (items)
            items = data.get("items", [])
            if isinstance(items, list) and len(items) > 0:
                for item in items:
                    row = [
                        stt,
                        res.filename,
                        res.document_type.value,
                        inv_no,
                        date,
                        seller,
                        buyer,
                        self._get_val(item, ["description", "product_description"]),
                        self._get_val(item, ["quantity"]),
                        self._get_val(item, ["unit"]),
                        self._get_val(item, ["unit_price"]),
                        self._get_val(item, ["amount"]),
                        currency,
                        incoterms,
                        f"{res.overall_confidence*100:.1f}"
                    ]
                    ws.append(row)
                    stt += 1
            else:
                # Nếu không có items (chứng từ đơn lẻ), xuất 1 dòng duy nhất
                row = [
                    stt, res.filename, res.document_type.value, inv_no, date,
                    seller, buyer, 
                    self._get_val(data, ["product_description", "description_of_goods"]),
                    self._get_val(data, ["quantity"]), "", "", 
                    self._get_val(data, ["total_amount"]),
                    currency, incoterms, f"{res.overall_confidence*100:.1f}"
                ]
                ws.append(row)
                stt += 1

        self._autofit_columns(ws)
        wb.save(output_path)
        logger.info("Excel Data Entry file saved at %s", output_path)

    def _get_val(self, data: dict, keys: list) -> str:
        """Lấy giá trị từ dict dựa trên danh sách các key ưu tiên."""
        for k in keys:
            if k in data:
                val = data[k]
                if isinstance(val, dict):
                    return str(val.get("value", ""))
                return str(val)
        return ""

    def _autofit_columns(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
